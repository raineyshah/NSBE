# -*- coding: utf-8 -*-
"""
Created on Mon Jan 27 15:42:09 2020

@author: rainey
"""

import openpyxl as xl 
from mechanize import Browser
from bs4 import BeautifulSoup 

wb = xl.load_workbook("master_attendance_spring.xlsx")
ws = wb['Sheet1']

for r in range(4,5):
    eid = ws.cell(row = r, column = 1).value 
    browser = Browser()
    browser.open("https://directory.utexas.edu/")
    browser.select_form(nr=0)
    browser['q'] = eid
    response = browser.submit()
    content = response.read()
    soup = BeautifulSoup(content, 'html.parser')
    
    page_text = soup.get_text().split('\n')
    i = 0
    while(1):
        try:
            if "Name:" in page_text[i]:
                i += 3
                name = page_text[i]
        except:
            break
        if "Email:" in page_text[i]:
            i += 3
            email = page_text[i]
        if "Major:" in page_text[i]:
            i += 3 
            major = page_text[i]
        if "Classification:" in page_text[i]:
            i += 3
            classification = page_text[i].split("/")[0]
            ws.cell(row = r, column = 2).value = name
            ws.cell(row = r, column = 3).value = email
            ws.cell(row = r, column = 4).value = major
            ws.cell(row = r, column = 5).value = classification
            break
        i += 1
wb.save("master_attendance_spring.xlsx")